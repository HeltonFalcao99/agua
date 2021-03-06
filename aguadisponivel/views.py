from django.shortcuts import render
from .forms import *
import openpyxl
from openpyxl import Workbook, load_workbook
from itertools import islice
from .models import PTFdados
from .forms import PTFform
from django.http import HttpResponse, Http404, HttpResponseRedirect
from os import path
from .PTF import Oliveira, Barros, Tomasella, BarrosSimplificada
import csv
import mimetypes
import matplotlib.pyplot as plt
from matplotlib.backends.backend_agg import FigureCanvasAgg as FigureCanvas
from matplotlib.figure import Figure


def inicio(request):
    return render(request, "aguadisponivel/home.html")



def excel(self, request):

    if "GET" == request.method:
        return render(request, 'aguadisponivel/excel.html', {})

    else:
        tabela = request.FILES["tabela"]
        wb = openpyxl.load_workbook(tabela)
        ws = wb.active
        #ws = ws.insert_cols( 15, amount = 1 )
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="respostaptfs.csv"'
        writer = csv.writer(response)
        writer.writerow(['OLIVEIRA', 'ALPHA BARROS', 'N BARROS', 'THETA R BARROS', 'THETA S BARROS', '0.01m3/m3 BARROS', '0.033m3/m3 BARROS', '1.5m3/m3 BARROS', 'Agua Disponivel BARROS',
                         'ALPHA TOMASELLA', 'N TOMASELLA', 'THETA R TOMASELLA', 'THETA S TOMASELLA', '0.01m3/m3 TOMASELLA', '0.033m3/m3 TOMASELLA', '1.5m3/m3 TOMASELLA', 'Agua Disponivel TOMASELLA'])
        for i in range(2, (ws.max_row + 1)):
            SBICS = ws.cell(row=i, column=1)  # SBICS
            if SBICS.value != None:
                SBICS = SBICS.value
            else:
                SBICS = None
            OLD = ws.cell(row=i, column=2)		# OLD
            if OLD.value != None:
                OLD = OLD.value
            else:
                OLD = None
            LAT = ws.cell(row=i, column=3)  # LAT
            if (LAT.value) != None:
                LAT = LAT.value
            else:
                LAT = None
            LON = ws.cell(row=i, column=4)  # LON
            if LON.value != None:
                LON = LON.value
            else:
                LON = None
            MUN = ws.cell(row=i, column=5)    	# MUN
            if MUN.value != None:
                MUN = MUN.value
            else:
                MUN = None
            GEOCOD = ws.cell(row=i, column=6)  # GEOCOD
            if GEOCOD.value != None:
                GEOCOD = GEOCOD.value
            else:
                GEOCOD = None
            T_ARE = ws.cell(row=i, column=7)  # T_ARE
            if T_ARE.value != None:
                T_ARE = T_ARE.value
            else:
                T_ARE = None
            AGRO = ws.cell(row=i, column=8)		# AGRO
            if AGRO.value != None:
                AGRO = AGRO.value
            else:
                AGRO = None
            AFIN = ws.cell(row=i, column=9)  # AFIN
            if AFIN.value != None:
                AFIN = AFIN.value
            else:
                AFIN = None
            SILT = ws.cell(row=i, column=10)  # SILT
            if SILT.value != None:
                SILT = SILT.value
            else:
                SILT = None
            ARG = ws.cell(row=i, column=11)  # ARG
            if ARG.value != None:
                ARG = ARG.value
            else:
                ARG = None
            SOLO = ws.cell(row=i, column=12)  # SOLO
            if SOLO.value != None:
                SOLO = float(SOLO.value)
            else:
                SOLO = None
            C_ORG = ws.cell(row=i, column=13)  # C_ORG
            if C_ORG.value != None:
                C_ORG = float(C_ORG.value)
                M_ORG = (2*C_ORG)
            else:
                C_ORG = None
                M_ORG = None

            if (T_ARE != None or SILT != None or ARG != None or SOLO != None or C_ORG != None and AGRO != None and AFIN != None and LAT != None and LON != None and MUN != None and GEOCOD != None):
                if((ARG + SILT + T_ARE) == 1000) and (SOLO <= 2) and (SOLO >= 0.8) and (type(LAT) == float) and (type(LON) == float):
                    dados = PTFdados(SBICS=SBICS, OLD=OLD, LAT=LAT, LON=LON, MUN=MUN, GEOCOD=GEOCOD, T_ARE=T_ARE, AGRO=AGRO,
                                     AFIN=AFIN, SILT=SILT, ARG=ARG, SOLO=SOLO, C_ORG=C_ORG, M_ORG=M_ORG)

                    oliv = Oliveira.oliveira(self ,T_ARE, SILT, ARG, SOLO)
                    b_alpha = Barros.bar_alpha(self ,T_ARE, ARG, SOLO)
                    b_n = Barros.bar_n(self ,T_ARE, SILT, M_ORG)
                    b_thetar = Barros.bar_thetar(self ,T_ARE, ARG, M_ORG, SOLO)
                    b_thetas = Barros.bar_thetas(self ,SOLO)
                    b_zu = Barros.bar_zu(self ,b_alpha, b_n, b_thetar, b_thetas)
                    b_tt = Barros.bar_zu(self ,b_alpha, b_n, b_thetar, b_thetas)
                    b_pc = Barros.bar_pc(self ,b_alpha, b_n, b_thetar, b_thetas)
                    b_ad = Barros.bar_ad(self ,b_tt, b_pc)
                    t_alpha = Tomasella.tom_alpha(self ,
                        (AGRO/10), (AFIN/10), (SILT/10), SOLO, (ARG/10))
                    t_n = Tomasella.tom_N(self ,
                        (AGRO/10), (AFIN/10), (SILT/10), (ARG/10))
                    t_thetar = Tomasella.tom_thetar(self ,
                        (AGRO/10), (AFIN/10), (SILT/10), SOLO, (ARG/10))
                    t_thetas = Tomasella.tom_thetas(self ,
                        (AGRO/10), (AFIN/10), (SILT/10), SOLO, C_ORG)
                    t_zu = Tomasella.tom_zu(self ,t_alpha, t_n, t_thetar, t_thetas)
                    t_tt = Tomasella.tom_tt(self ,t_alpha, t_n, t_thetar, t_thetas)
                    t_pc = Tomasella.tom_pc(self ,t_alpha, t_n, t_thetar, t_thetas)
                    t_ad = Tomasella.tom_ad(self ,t_tt, t_pc)

                    dados.oliv = oliv
                    dados.B_ZU = b_zu
                    dados.B_TT = b_tt
                    dados.B_PC = b_pc
                    dados.B_ALPHA = b_alpha
                    dados.B_N = b_n
                    dados.B_THETAS = b_thetas
                    dados.B_TETHAR = b_thetar
                    dados.B_AD = b_ad
                    dados.T_ZU = t_zu
                    dados.T_TT = t_tt
                    dados.T_PC = t_pc
                    dados.T_ALPHA = t_alpha
                    dados.T_N = t_n
                    dados.T_THETAR = t_thetar
                    dados.T_TETHAS = t_thetas
                    dados.T_AD = t_ad
                    dados.save()

                    oliv = str(oliv)
                    b_alpha = str(b_alpha)
                    b_n = str(b_n)
                    b_thetar = str(b_thetar)
                    b_thetas = str(b_thetas)
                    t_alpha = str(t_alpha)
                    t_n = str(t_n)
                    t_thetar = str(t_thetar)
                    t_thetas = str(t_thetas)
                    b_zu = str(b_zu)
                    b_tt = str(b_tt)
                    b_pc = str(b_pc)
                    t_zu = str(t_zu)
                    t_tt = str(t_tt)
                    t_pc = str(t_pc)
                    b_ad = str(b_ad)
                    t_ad = str(t_ad)

                    writer.writerow([oliv, b_alpha, b_n, b_thetar, b_thetas, b_zu, b_tt,
                                     b_pc, b_ad, t_alpha, t_n, t_thetar, t_thetas, t_zu, t_tt, t_pc, t_ad])
        return response
    return render(request, 'aguadisponivel/excel.html')

  
def PTFview(self ,request):
    form = PTFform(request.POST)
    if "GET" == request.method:
        oliv = '', b_alpha = '', b_n = '', b_thetar = '', b_thetas = '', t_alpha = '', t_n = '', t_thetar = '', t_thetas = '', b_zu = '',
        b_tt = '', b_pc = '', t_zu = '', t_tt = '', t_pc = '', b_ad = '', bs_logalpha = '', bs_alpha = '', bs_n = 'Na', bs_thetar = '',
        bs_thetas = '', bs_zu = '', bs_tt = '', bs_pc = '', bs_ad = '', t_ad = 'Na', ad_cliente = ''

        return render(request, "aguadisponivel/PTFtemplate.html", {'form': form, 'oliv': oliv, 'b_alpha': b_alpha, 'b_n': b_n, 'b_thetar': b_thetar,
        'b_thetas': b_thetas, 'b_zu': b_zu, 'b_tt': b_tt, 'b_pc': b_pc, 'b_ad': b_ad, 'bs_logalpha': bs_logalpha, 'bs_alpha': bs_alpha, 'bs_n': bs_n,
         'bs_thetar': bs_thetar, 'bs_thetas': bs_thetas, 'bs_zu': bs_zu, 'bs_tt': bs_tt, 'bs_pc': bs_pc, 'bs_ad': bs_ad, 't_alpha': t_alpha, 't_n': t_n,
          't_thetar': t_thetar, 't_thetas': t_thetas, 't_zu': t_zu, 't_tt': t_tt, 't_pc': t_pc, 't_ad': t_ad, 'ad_cliente': ad_cliente})

    elif request.method == "POST":
        oliv = '', b_alpha = '', b_n = '', b_thetar = '', b_thetas = '', t_alpha = '', t_n = '', t_thetar = '', t_thetas = '', b_zu = '',
        b_tt = '', b_pc = '', t_zu = '', t_tt = '', t_pc = '', b_ad = '', bs_logalpha = '', bs_alpha = '', bs_n = '', bs_thetar = '',
        bs_thetas = '', bs_zu = '', bs_tt = '', bs_pc = '', bs_ad = '', t_ad = 'Na', ad_cliente = ''

    if form.is_valid():

            SBICS = form.cleaned_data['SBICS'], OLD = form.cleaned_data["OLD"], LAT = form.cleaned_data["LAT"], LON = form.cleaned_data["LON"]
            MUN = form.cleaned_data['MUN'], GEOCOD = form.cleaned_data['GEOCOD'], T_ARE = form.cleaned_data['T_ARE'], AGRO = form.cleaned_data['AGRO']
            AFIN = form.cleaned_data['AFIN'], SILT = form.cleaned_data['SILT'], ARG = form.cleaned_data['ARG'], SOLO = form.cleaned_data['SOLO']
            C_ORG = form.cleaned_data['C_ORG']
            if(C_ORG != None):
                M_ORG = (2*C_ORG)
            B_TT = form.cleaned_data['B_TT'], B_PC = form.cleaned_data['B_PC']


        if((ARG + SILT + T_ARE) == 1000) and (SOLO <= 2) and (SOLO >= 0.8) and (T_ARE == (AFIN + AGRO)) and (C_ORG != ''):
                    oliv = Oliveira.oliveira(self ,T_ARE, SILT, ARG, SOLO)

                    bs_logalpha = BarrosSimplificada.BS_logalpha(self ,ARG)
                    bs_alpha = BarrosSimplificada.BS_alpha(self ,bs_logalpha)
                    bs_n = BarrosSimplificada.BS_n(self ,T_ARE, SILT)
                    bs_thetar = BarrosSimplificada.BS_thetar(self ,T_ARE, ARG)
                    bs_thetas = BarrosSimplificada.BS_thetas(self ,T_ARE, SILT)
                    bs_zu = BarrosSimplificada.BS_zu(self ,
                        bs_alpha, bs_n, bs_thetar, bs_thetas)
                    bs_tt = BarrosSimplificada.BS_tt(self ,
                        bs_alpha, bs_n, bs_thetar, bs_thetas)
                    bs_pc = BarrosSimplificada.BS_pc(self ,
                        bs_alpha, bs_n, bs_thetar, bs_thetas)

                    b_alpha = Barros.bar_alpha(self ,T_ARE, ARG, SOLO)
                    b_n = Barros.bar_n(self ,T_ARE, SILT, M_ORG)
                    b_thetar = Barros.bar_thetar(self ,T_ARE, ARG, M_ORG, SOLO)
                    b_thetas = Barros.bar_thetas(self ,SOLO)
                    b_zu = Barros.bar_zu(self ,b_alpha, b_n, b_thetar, b_thetas)
                    b_tt = Barros.bar_tt(self ,b_alpha, b_n, b_thetar, b_thetas)
                    b_pc = Barros.bar_pc(self ,b_alpha, b_n, b_thetar, b_thetas)
                    b_ad = Barros.bar_ad(self ,b_tt, b_pc)

                    t_alpha = Tomasella.tom_alpha(self ,
                        (AGRO/10), (AFIN/10), (SILT/10), SOLO, (ARG/10))
                    t_n = Tomasella.tom_N(self ,
                        (AGRO/10), (AFIN/10), (SILT/10), (ARG/10))
                    t_thetar = Tomasella.tom_thetar(self ,
                        (AGRO/10), (AFIN/10), (SILT/10), SOLO, (ARG/10))
                    t_thetas = Tomasella.tom_thetas(self ,
                        (AGRO/10), (AFIN/10), (SILT/10), SOLO, C_ORG)
                    t_zu = Tomasella.tom_zu(self ,t_alpha, t_n, t_thetar, t_thetas)
                    t_tt = Tomasella.tom_tt(self ,t_alpha, t_n, t_thetar, t_thetas)
                    t_pc = Tomasella.tom_pc(self ,t_alpha, t_n, t_thetar, t_thetas)
                    t_ad = Tomasella.tom_ad(self ,t_tt, t_pc)

                    return render(request, "aguadisponivel/PTFtemplate.html", {'form': form, 'oliv': oliv, 'b_alpha': b_alpha, 'b_n': b_n, 'b_thetar': b_thetar, 'b_thetas': b_thetas, 'b_zu': b_zu, 'b_tt': b_tt, 'b_pc': b_pc, 'b_ad': b_ad, 'bs_logalpha': bs_logalpha, 'bs_alpha': bs_alpha, 'bs_n': bs_n, 'bs_thetar': bs_thetar, 'bs_thetas': bs_thetas, 'bs_zu': bs_zu, 'bs_tt': bs_tt, 'bs_pc': bs_pc, 'bs_ad': bs_ad, 't_alpha': t_alpha, 't_n': t_n, 't_thetar': t_thetar, 't_thetas': t_thetas, 't_zu': t_zu, 't_tt': t_tt, 't_pc': t_pc, 't_ad': t_ad, 'ad_cliente': ad_cliente})


"""
def PTFview(self ,request):
    form = PTFform(request.POST)
    if "GET" == request.method:
        oliv = ''
        b_alpha = ''
        b_n = ''
        b_thetar = ''
        b_thetas = ''
        t_alpha = ''
        t_n = ''
        t_thetar = ''
        t_thetas = ''
        b_zu = ''
        b_tt = ''
        b_pc = ''
        t_zu = ''
        t_tt = ''
        t_pc = ''
        b_ad = ''
        bs_logalpha = ''
        bs_alpha = ''
        bs_n = 'Na'
        bs_thetar = ''
        bs_thetas = ''
        bs_zu = ''
        bs_tt = ''
        bs_pc = ''
        bs_ad = ''
        escolha = ''
        t_ad = 'Na'
        ad_cliente = ''

        return render(request, "aguadisponivel/PTFtemplate.html", {'form': form, 'oliv': oliv, 'b_alpha': b_alpha, 'b_n': b_n, 'b_thetar': b_thetar, 'b_thetas': b_thetas, 'b_zu': b_zu, 'b_tt': b_tt, 'b_pc': b_pc, 'b_ad': b_ad, 'bs_logalpha': bs_logalpha, 'bs_alpha': bs_alpha, 'bs_n': bs_n, 'bs_thetar': bs_thetar, 'bs_thetas': bs_thetas, 'bs_zu': bs_zu, 'bs_tt': bs_tt, 'bs_pc': bs_pc, 'bs_ad': bs_ad, 't_alpha': t_alpha, 't_n': t_n, 't_thetar': t_thetar, 't_thetas': t_thetas, 't_zu': t_zu, 't_tt': t_tt, 't_pc': t_pc, 't_ad': t_ad, 'ad_cliente': ad_cliente})
    elif request.method == "POST":
        oliv = ''
        b_alpha = ''
        b_n = ''
        b_thetar = ''
        b_thetas = ''
        t_alpha = ''
        t_n = ''
        t_thetar = ''
        t_thetas = ''
        b_zu = ''
        b_tt = ''
        b_pc = ''
        t_zu = ''
        t_tt = ''
        t_pc = ''
        b_ad = ''
        bs_logalpha = ''
        bs_alpha = ''
        bs_n = ''
        bs_thetar = ''
        bs_thetas = ''
        bs_zu = ''
        bs_tt = ''
        bs_pc = ''
        bs_ad = ''
        escolha = ''
        t_ad = 'Na'
        ad_cliente = ''
        
        if form.is_valid():

            escolha = form.cleaned_data["escolha"]
            SBICS = form.cleaned_data['SBICS']
            OLD = form.cleaned_data["OLD"]
            LAT = form.cleaned_data["LAT"]
            LON = form.cleaned_data["LON"]
            MUN = form.cleaned_data['MUN']
            GEOCOD = form.cleaned_data['GEOCOD']
            T_ARE = form.cleaned_data['T_ARE']
            AGRO = form.cleaned_data['AGRO']
            AFIN = form.cleaned_data['AFIN']
            SILT = form.cleaned_data['SILT']
            ARG = form.cleaned_data['ARG']
            SOLO = form.cleaned_data['SOLO']
            C_ORG = form.cleaned_data['C_ORG']
            if(C_ORG != None):
                M_ORG = (2*C_ORG)
            B_TT = form.cleaned_data['B_TT']
            B_PC = form.cleaned_data['B_PC']

            # caso seja so tipo de solo
            if(escolha == 'ct'):

                return render(request, "aguadisponivel/PTFtemplate.html", {'form': form, 'oliv': oliv, 'b_alpha': b_alpha, 'b_n': b_n, 'b_thetar': b_thetar, 'b_thetas': b_thetas, 'b_zu': b_zu, 'b_tt': b_tt, 'b_pc': b_pc, 'b_ad': b_ad, 'bs_logalpha': bs_logalpha, 'bs_alpha': bs_alpha, 'bs_n': bs_n, 'bs_thetar': bs_thetar, 'bs_thetas': bs_thetas, 'bs_zu': bs_zu, 'bs_tt': bs_tt, 'bs_pc': bs_pc, 'bs_ad': bs_ad, 't_alpha': t_alpha, 't_n': t_n, 't_thetar': t_thetar, 't_thetas': t_thetas, 't_zu': t_zu, 't_tt': t_tt, 't_pc': t_pc, 't_ad': t_ad, 'ad_cliente': ad_cliente})

            elif(escolha == 'asa'):
                if((ARG + SILT + T_ARE) == 1000):

                    bs_logalpha = BarrosSimplificada.BS_logalpha(self ,ARG)
                    bs_alpha = BarrosSimplificada.BS_alpha(self ,bs_logalpha)
                    bs_n = BarrosSimplificada.BS_n(self ,T_ARE, SILT)
                    bs_thetar = BarrosSimplificada.BS_thetar(self ,T_ARE, ARG)
                    bs_thetas = BarrosSimplificada.BS_thetas(self ,T_ARE, SILT)
                    bs_zu = BarrosSimplificada.BS_zu(self ,bs_alpha, bs_n, bs_thetar, bs_thetas)
                    bs_tt = BarrosSimplificada.BS_tt(self ,bs_alpha, bs_n, bs_thetar, bs_thetas)
                    bs_pc = BarrosSimplificada.BS_pc(self ,bs_alpha, bs_n, bs_thetar, bs_thetas)

                return render(request, "aguadisponivel/PTFtemplate.html", {'form': form, 'oliv': oliv, 'b_alpha': b_alpha, 'b_n': b_n, 'b_thetar': b_thetar, 'b_thetas': b_thetas, 'b_zu': b_zu, 'b_tt': b_tt, 'b_pc': b_pc, 'b_ad': b_ad, 'bs_logalpha': bs_logalpha, 'bs_alpha': bs_alpha, 'bs_n': bs_n, 'bs_thetar': bs_thetar, 'bs_thetas': bs_thetas, 'bs_zu': bs_zu, 'bs_tt': bs_tt, 'bs_pc': bs_pc, 'bs_ad': bs_ad, 't_alpha': t_alpha, 't_n': t_n, 't_thetar': t_thetar, 't_thetas': t_thetas, 't_zu': t_zu, 't_tt': t_tt, 't_pc': t_pc, 't_ad': t_ad, 'ad_cliente': ad_cliente})

            elif(escolha == 'asad'):
                if((ARG + SILT + T_ARE) == 1000) and (SOLO <= 2) and (SOLO >= 0.8) and (T_ARE == (AFIN + AGRO)) and (C_ORG != ''):
                    oliv = Oliveira.oliveira(self ,T_ARE, SILT, ARG, SOLO)

                    bs_logalpha = BarrosSimplificada.BS_logalpha(self ,ARG)
                    bs_alpha = BarrosSimplificada.BS_alpha(self ,bs_logalpha)
                    bs_n = BarrosSimplificada.BS_n(self ,T_ARE, SILT)
                    bs_thetar = BarrosSimplificada.BS_thetar(self ,T_ARE, ARG)
                    bs_thetas = BarrosSimplificada.BS_thetas(self ,T_ARE, SILT)
                    bs_zu = BarrosSimplificada.BS_zu(self ,
                        bs_alpha, bs_n, bs_thetar, bs_thetas)
                    bs_tt = BarrosSimplificada.BS_tt(self ,
                        bs_alpha, bs_n, bs_thetar, bs_thetas)
                    bs_pc = BarrosSimplificada.BS_pc(self ,
                        bs_alpha, bs_n, bs_thetar, bs_thetas)

                    b_alpha = Barros.bar_alpha(self ,T_ARE, ARG, SOLO)
                    b_n = Barros.bar_n(self ,T_ARE, SILT, M_ORG)
                    b_thetar = Barros.bar_thetar(self ,T_ARE, ARG, M_ORG, SOLO)
                    b_thetas = Barros.bar_thetas(self ,SOLO)
                    b_zu = Barros.bar_zu(self ,b_alpha, b_n, b_thetar, b_thetas)
                    b_tt = Barros.bar_tt(self ,b_alpha, b_n, b_thetar, b_thetas)
                    b_pc = Barros.bar_pc(self ,b_alpha, b_n, b_thetar, b_thetas)
                    b_ad = Barros.bar_ad(self ,b_tt, b_pc)

                    t_alpha = Tomasella.tom_alpha(self ,
                        (AGRO/10), (AFIN/10), (SILT/10), SOLO, (ARG/10))
                    t_n = Tomasella.tom_N(self ,
                        (AGRO/10), (AFIN/10), (SILT/10), (ARG/10))
                    t_thetar = Tomasella.tom_thetar(self ,
                        (AGRO/10), (AFIN/10), (SILT/10), SOLO, (ARG/10))
                    t_thetas = Tomasella.tom_thetas(self ,
                        (AGRO/10), (AFIN/10), (SILT/10), SOLO, C_ORG)
                    t_zu = Tomasella.tom_zu(self ,t_alpha, t_n, t_thetar, t_thetas)
                    t_tt = Tomasella.tom_tt(self ,t_alpha, t_n, t_thetar, t_thetas)
                    t_pc = Tomasella.tom_pc(self ,t_alpha, t_n, t_thetar, t_thetas)
                    t_ad = Tomasella.tom_ad(self ,t_tt, t_pc)

                    return render(request, "aguadisponivel/PTFtemplate.html", {'form': form, 'oliv': oliv, 'b_alpha': b_alpha, 'b_n': b_n, 'b_thetar': b_thetar, 'b_thetas': b_thetas, 'b_zu': b_zu, 'b_tt': b_tt, 'b_pc': b_pc, 'b_ad': b_ad, 'bs_logalpha': bs_logalpha, 'bs_alpha': bs_alpha, 'bs_n': bs_n, 'bs_thetar': bs_thetar, 'bs_thetas': bs_thetas, 'bs_zu': bs_zu, 'bs_tt': bs_tt, 'bs_pc': bs_pc, 'bs_ad': bs_ad, 't_alpha': t_alpha, 't_n': t_n, 't_thetar': t_thetar, 't_thetas': t_thetas, 't_zu': t_zu, 't_tt': t_tt, 't_pc': t_pc, 't_ad': t_ad, 'ad_cliente': ad_cliente})


            elif(escolha == 'assadtt'):
                return ''

            elif(escolha == 'assadpc'):
                

                # if (T_ARE != None and SILT != None and ARG != None and SOLO != None and C_ORG != None and AGRO != None and AFIN != None and LAT != None and LON != None and MUN != None and GEOCOD != None):
                # and (T_ARE == (AFIN + AGRO)) and (type(LAT) == float) and (type(LON) == float):
                if((ARG + SILT + T_ARE) == 1000) and (SOLO <= 2) and (SOLO >= 0.8):

                    oliv = Oliveira.oliveira(self ,T_ARE, SILT, ARG, SOLO)

                    bs_logalpha = BarrosSimplificada.BS_logalpha(self ,ARG)
                    bs_alpha = BarrosSimplificada.BS_alpha(self ,bs_logalpha)
                    bs_n = BarrosSimplificada.BS_n(self ,T_ARE, SILT)
                    bs_thetar = BarrosSimplificada.BS_thetar(self ,T_ARE, ARG)
                    bs_thetas = BarrosSimplificada.BS_thetas(self ,T_ARE, SILT)
                    bs_zu = BarrosSimplificada.BS_zu(self ,
                        bs_alpha, bs_n, bs_thetar, bs_thetas)
                    bs_tt = BarrosSimplificada.BS_tt(self ,
                        bs_alpha, bs_n, bs_thetar, bs_thetas)
                    bs_pc = BarrosSimplificada.BS_pc(self ,
                        bs_alpha, bs_n, bs_thetar, bs_thetas)

                    b_alpha = Barros.bar_alpha(self ,T_ARE, ARG, SOLO)
                    b_n = Barros.bar_n(self ,T_ARE, SILT, M_ORG)
                    b_thetar = Barros.bar_thetar(self ,T_ARE, ARG, M_ORG, SOLO)
                    b_thetas = Barros.bar_thetas(self ,SOLO)
                    b_zu = Barros.bar_zu(self ,b_alpha, b_n, b_thetar, b_thetas)
                    b_tt = Barros.bar_tt(self ,b_alpha, b_n, b_thetar, b_thetas)
                    b_pc = Barros.bar_pc(self ,b_alpha, b_n, b_thetar, b_thetas)
                    b_ad = Barros.bar_ad(self ,b_tt, b_pc)

                    t_alpha = Tomasella.tom_alpha(self ,
                        (AGRO/10), (AFIN/10), (SILT/10), SOLO, (ARG/10))
                    t_n = Tomasella.tom_N(self ,
                        (AGRO/10), (AFIN/10), (SILT/10), (ARG/10))
                    t_thetar = Tomasella.tom_thetar(self ,
                        (AGRO/10), (AFIN/10), (SILT/10), SOLO, (ARG/10))
                    t_thetas = Tomasella.tom_thetas(self ,
                        (AGRO/10), (AFIN/10), (SILT/10), SOLO, C_ORG)
                    t_zu = Tomasella.tom_zu(self ,t_alpha, t_n, t_thetar, t_thetas)
                    t_tt = Tomasella.tom_tt(self ,t_alpha, t_n, t_thetar, t_thetas)
                    t_pc = Tomasella.tom_pc(self ,t_alpha, t_n, t_thetar, t_thetas)
                    t_ad = Tomasella.tom_ad(self ,t_tt, t_pc)
                    ad_cliente = Tomasella.tom_ad(self ,B_TT,B_PC)

                if((LAT != '' and LON != '') or (MUN != '')):

                    dados = PTFdados(SBICS=SBICS, OLD=OLD, LAT=LAT, LON=LON, MUN=MUN, GEOCOD=GEOCOD, T_ARE=T_ARE,
                                    AGRO=AGRO, AFIN=AFIN, SILT=SILT, ARG=ARG, SOLO=SOLO, C_ORG=C_ORG, M_ORG=M_ORG)
                    dados = PTFdados(oliv=oliv, B_ZU=b_zu, B_TT=b_tt, B_PC=b_pc, B_ALPHA=b_alpha, B_N=b_n, B_THETAS=b_thetas, B_TETHAR=b_thetar,
                                    B_AD=b_ad, T_ZU=t_zu, T_TT=t_tt, T_PC=t_pc, T_ALPHA=t_alpha, T_N=t_n, T_THETAR=t_thetar, T_TETHAS=t_thetas, T_AD=t_ad)
                    dados.save()
                
                oliv = str(oliv)
                b_alpha = str(b_alpha)
                b_n = str(b_n)
                b_thetar = str(b_thetar)
                b_thetas = str(b_thetas)
                t_alpha = str(t_alpha)
                t_n = str(t_n)
                t_thetar = str(t_thetar)
                t_thetas = str(t_thetas)
                b_zu = str(b_zu)
                b_tt = str(b_tt)
                b_pc = str(b_pc)
                t_zu = str(t_zu)
                t_tt = str(t_tt)
                t_pc = str(t_pc)
                b_ad = str(b_ad)
                t_ad = str(t_ad)
                ad_cliente = str(ad_cliente)
                return render(request, "aguadisponivel/PTFtemplate.html", {'form': form, 'oliv': oliv, 'b_alpha': b_alpha, 'b_n': b_n, 'b_thetar': b_thetar, 'b_thetas': b_thetas, 'b_zu': b_zu, 'b_tt': b_tt, 'b_pc': b_pc, 'b_ad': b_ad, 'bs_logalpha': bs_logalpha, 'bs_alpha': bs_alpha, 'bs_n': bs_n, 'bs_thetar': bs_thetar, 'bs_thetas': bs_thetas, 'bs_zu': bs_zu, 'bs_tt': bs_tt, 'bs_pc': bs_pc, 'bs_ad': bs_ad, 't_alpha': t_alpha, 't_n': t_n, 't_thetar': t_thetar, 't_thetas': t_thetas, 't_zu': t_zu, 't_tt': t_tt, 't_pc': t_pc, 't_ad': t_ad, 'ad_cliente': ad_cliente})

        else:
            return render(request, "aguadisponivel/home.html")


""" 